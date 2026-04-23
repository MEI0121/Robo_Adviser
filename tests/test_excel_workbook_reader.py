"""
Unit tests for the reconcile.read_excel_reconciliation_data reader.

The reader extracts reconciliation values from known cell locations in
the team's Excel audit workbook. Two behaviours must be pinned:

  1. When cells contain cached values (i.e. the workbook was opened in
     Excel and saved, or a test wrote values directly via openpyxl),
     the reader returns numpy arrays with the correct shape, dtype,
     and values.
  2. When cells are empty (freshly-built openpyxl workbook with formulas
     but no cached values; happens on every xlsx that has never been
     opened in Microsoft Excel), the reader returns None for that
     key — enabling the per-check CSV fallback downstream.

The fixture workbook is built in a fixture rather than shipped as a
binary asset so the known values are transparently reviewable in code.
"""

from __future__ import annotations

import os
import sys
from pathlib import Path

import numpy as np
import pytest

# reconcile.py lives at the project root; tests/ is one level down
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "backend"))

import reconcile  # noqa: E402


# ---------------------------------------------------------------------------
# Fixture workbook — pre-baked with known cached values
# ---------------------------------------------------------------------------


def _build_test_workbook(path: Path) -> dict:
    """
    Build a minimal 12-sheet workbook with known values in every cell the
    reader reads. Returns a dict of the expected values, keyed the same
    way the reader's output should be.
    """
    from openpyxl import Workbook  # noqa: PLC0415

    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)
    for name in (
        "README", "NAV_Data", "Log_Returns", "Cov_Matrix", "GMVP", "Optimal",
        "Frontier", "Frontier_Chart", "Equal_Weight", "Tangency",
        "GMVP_Short", "Frontier_Short",
    ):
        wb.create_sheet(name)

    expected: dict = {}

    # Cov_Matrix!B15:K15 — mu (10-element row)
    mu = np.linspace(0.03, 0.18, 10)
    ws = wb["Cov_Matrix"]
    for i, v in enumerate(mu):
        ws.cell(row=15, column=2 + i, value=float(v))
    expected["mu"] = mu

    # Cov_Matrix!B2:K11 — covariance (10x10)
    rng = np.random.default_rng(0)
    A = rng.standard_normal((10, 10)) * 0.05
    cov = A @ A.T + np.eye(10) * 0.02
    for r in range(10):
        for c in range(10):
            ws.cell(row=2 + r, column=2 + c, value=float(cov[r, c]))
    expected["cov"] = cov

    # GMVP!B27:B39 — 10 weights + 3 stats
    gmvp_w = np.array([0.02, 0.00, 0.02, 0.00, 0.00, 0.00, 0.00, 0.96, 0.00, 0.00])
    gmvp_s = np.array([0.02643, 0.04715, -0.07574])
    ws = wb["GMVP"]
    for i, v in enumerate(gmvp_w):
        ws.cell(row=27 + i, column=2, value=float(v))
    for i, v in enumerate(gmvp_s):
        ws.cell(row=37 + i, column=2, value=float(v))
    expected["gmvp_weights"] = gmvp_w
    expected["gmvp_stats"] = gmvp_s

    # GMVP_Short!B30:B42
    gs_w = np.array([-0.13, 0.08, 0.08, 0.36, -0.18, -0.18, 0.04, 0.99, 0.02, -0.08])
    gs_s = np.array([0.02383, 0.04100, -0.15040])
    ws = wb["GMVP_Short"]
    for i, v in enumerate(gs_w):
        ws.cell(row=30 + i, column=2, value=float(v))
    for i, v in enumerate(gs_s):
        ws.cell(row=40 + i, column=2, value=float(v))
    expected["gmvp_short_weights"] = gs_w
    expected["gmvp_short_stats"] = gs_s

    # Equal_Weight!B27:B39 (weights all 0.1, stats are placeholders)
    ew_w = np.full(10, 0.1)
    ew_s = np.array([0.09270, 0.11890, 0.52750])
    ws = wb["Equal_Weight"]
    for i, v in enumerate(ew_w):
        ws.cell(row=27 + i, column=2, value=float(v))
    for i, v in enumerate(ew_s):
        ws.cell(row=37 + i, column=2, value=float(v))
    expected["equal_weight_weights"] = ew_w
    expected["equal_weight_stats"] = ew_s

    # Tangency!B48:B60
    t_w = np.array([0.0, 0.0, 0.2, 0.4, 0.0, 0.4, 0.0, 0.0, 0.0, 0.0])
    t_s = np.array([0.14630, 0.14760, 0.78770])
    ws = wb["Tangency"]
    for i, v in enumerate(t_w):
        ws.cell(row=48 + i, column=2, value=float(v))
    for i, v in enumerate(t_s):
        ws.cell(row=58 + i, column=2, value=float(v))
    expected["tangency_weights"] = t_w
    expected["tangency_stats"] = t_s

    # Frontier!B4:N103 — 100 rows × 13 cols: [target_return, vol, sharpe, w0..w9]
    frontier = rng.uniform(0, 0.3, (100, 13))
    ws = wb["Frontier"]
    for r in range(100):
        for c in range(13):
            ws.cell(row=4 + r, column=2 + c, value=float(frontier[r, c]))
    expected["frontier"] = frontier

    # Frontier_Short!B5:N104 — same shape, data starts one row lower
    frontier_s = rng.uniform(-0.2, 0.5, (100, 13))
    ws = wb["Frontier_Short"]
    for r in range(100):
        for c in range(13):
            ws.cell(row=5 + r, column=2 + c, value=float(frontier_s[r, c]))
    expected["frontier_short"] = frontier_s

    wb.save(path)
    return expected


@pytest.fixture(scope="module")
def populated_workbook(tmp_path_factory) -> tuple[Path, dict]:
    """Build a populated test workbook once per module."""
    path = tmp_path_factory.mktemp("wb") / "minimal_reconciliation.xlsx"
    expected = _build_test_workbook(path)
    return path, expected


# ---------------------------------------------------------------------------
# Behaviour 1: populated cells round-trip correctly through the reader
# ---------------------------------------------------------------------------


class TestReaderOnPopulatedWorkbook:
    def test_all_expected_keys_present(self, populated_workbook):
        path, expected = populated_workbook
        out = reconcile.read_excel_reconciliation_data(path)
        for key in expected:
            assert key in out, f"reader missing key {key!r} (got {sorted(out)})"

    def test_mu_roundtrips(self, populated_workbook):
        path, expected = populated_workbook
        out = reconcile.read_excel_reconciliation_data(path)
        assert out["mu"] is not None
        assert out["mu"].shape == (10,)
        np.testing.assert_allclose(out["mu"], expected["mu"], rtol=0, atol=1e-12)

    def test_cov_roundtrips(self, populated_workbook):
        path, expected = populated_workbook
        out = reconcile.read_excel_reconciliation_data(path)
        assert out["cov"] is not None
        assert out["cov"].shape == (10, 10)
        np.testing.assert_allclose(out["cov"], expected["cov"], rtol=0, atol=1e-12)

    @pytest.mark.parametrize(
        "key",
        [
            "gmvp_weights", "gmvp_stats",
            "gmvp_short_weights", "gmvp_short_stats",
            "equal_weight_weights", "equal_weight_stats",
            "tangency_weights", "tangency_stats",
        ],
    )
    def test_portfolio_block_roundtrips(self, populated_workbook, key):
        path, expected = populated_workbook
        out = reconcile.read_excel_reconciliation_data(path)
        assert out[key] is not None, f"{key} came back as None despite populated fixture"
        np.testing.assert_allclose(out[key], expected[key], rtol=0, atol=1e-12)

    @pytest.mark.parametrize("key", ["frontier", "frontier_short"])
    def test_frontier_shapes(self, populated_workbook, key):
        path, expected = populated_workbook
        out = reconcile.read_excel_reconciliation_data(path)
        assert out[key] is not None
        assert out[key].shape == (100, 13)
        np.testing.assert_allclose(out[key], expected[key], rtol=0, atol=1e-12)


# ---------------------------------------------------------------------------
# Behaviour 2: empty-cell cells return None (not zero, not exception)
# ---------------------------------------------------------------------------


class TestReaderOnEmptyWorkbook:
    """
    A workbook built by openpyxl with only formula cells — and no cached
    values because Microsoft Excel has never opened it — is exactly the
    state of the team's Group_BMD5302_Robo.xlsx before it's been opened
    and saved. The reader must return None for every key (not zero, not
    crash) so the per-check fallback to CSV fires downstream.
    """

    def test_empty_workbook_returns_all_none(self, tmp_path):
        from openpyxl import Workbook  # noqa: PLC0415

        wb = Workbook()
        wb.remove(wb.active)
        # Create only the sheets; leave them empty
        for name in (
            "Cov_Matrix", "GMVP", "GMVP_Short", "Equal_Weight",
            "Tangency", "Frontier", "Frontier_Short",
        ):
            wb.create_sheet(name)
        path = tmp_path / "empty.xlsx"
        wb.save(path)

        out = reconcile.read_excel_reconciliation_data(path)

        # Every key that the cell map produces should be None on an empty workbook
        for entry in reconcile._RECONCILIATION_CELL_MAP:
            key = entry["key"]
            keys = key if isinstance(key, tuple) else (key,)
            for k in keys:
                assert out.get(k) is None, (
                    f"empty workbook should yield None for {k}; got {out.get(k)!r}"
                )

    def test_missing_sheet_does_not_raise(self, tmp_path):
        """
        If a sheet the reader expects is missing entirely, the reader
        must return None for that key (not raise KeyError).
        """
        from openpyxl import Workbook  # noqa: PLC0415

        wb = Workbook()
        wb.remove(wb.active)
        # Only create ONE sheet — everything else is missing
        wb.create_sheet("Cov_Matrix")
        path = tmp_path / "one_sheet.xlsx"
        wb.save(path)

        # Should not raise
        out = reconcile.read_excel_reconciliation_data(path)

        # The keys whose sheets are missing should be None
        for key in (
            "gmvp_weights", "gmvp_stats",
            "gmvp_short_weights", "equal_weight_stats",
            "tangency_weights", "frontier", "frontier_short",
        ):
            assert out.get(key) is None, (
                f"expected None for {key} (its sheet is missing); got {out.get(key)!r}"
            )


# ---------------------------------------------------------------------------
# Behaviour 3: _find_excel_workbook preference order
# ---------------------------------------------------------------------------


class TestWorkbookDiscovery:
    def test_xlsm_preferred_over_xlsx(self, tmp_path):
        (tmp_path / "Group_BMD5302_Robo.xlsx").write_bytes(b"not a real xlsx")
        (tmp_path / "Group_BMD5302_Robo.xlsm").write_bytes(b"not a real xlsm")
        found = reconcile._find_excel_workbook(tmp_path)
        assert found is not None
        assert found.name.endswith(".xlsm")

    def test_xlsx_used_when_xlsm_absent(self, tmp_path):
        (tmp_path / "Group_BMD5302_Robo.xlsx").write_bytes(b"not a real xlsx")
        found = reconcile._find_excel_workbook(tmp_path)
        assert found is not None
        assert found.name.endswith(".xlsx")

    def test_returns_none_when_neither_present(self, tmp_path):
        assert reconcile._find_excel_workbook(tmp_path) is None


# ---------------------------------------------------------------------------
# Behaviour 4: _require_reconciliation_source prefers workbook over CSV
# ---------------------------------------------------------------------------


class TestSourcePreference:
    def test_workbook_value_preferred_when_populated(self, tmp_path):
        # Create a real CSV on disk with different values
        import pandas as pd  # noqa: PLC0415

        csv_path = tmp_path / "csv_source.csv"
        pd.DataFrame([[1.0], [2.0], [3.0]]).to_csv(csv_path, index=False, header=False)

        workbook_data = {"mu": np.array([9.0, 9.0, 9.0])}

        arr = reconcile._require_reconciliation_source(
            csv_path, "mu",
            workbook_data=workbook_data, workbook_key="mu",
        )
        np.testing.assert_array_equal(arr, [9.0, 9.0, 9.0])

    def test_csv_fallback_when_workbook_key_missing(self, tmp_path):
        import pandas as pd  # noqa: PLC0415

        csv_path = tmp_path / "csv_source.csv"
        pd.DataFrame([[1.0], [2.0], [3.0]]).to_csv(csv_path, index=False, header=False)

        workbook_data = {"mu": None}  # populated dict, this key is None

        arr = reconcile._require_reconciliation_source(
            csv_path, "mu",
            workbook_data=workbook_data, workbook_key="mu",
        )
        np.testing.assert_array_equal(arr.flatten(), [1.0, 2.0, 3.0])

    def test_skip_when_neither_available(self, tmp_path):
        arr = reconcile._require_reconciliation_source(
            tmp_path / "missing.csv", "mu",
            workbook_data={"mu": None}, workbook_key="mu",
        )
        assert arr is None
