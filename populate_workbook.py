"""
Populate the Excel workbook NAV_Data sheet from the project's raw CSV files.

Usage (from the repo root):
    python populate_workbook.py

Reads:
    data/raw/{TICKER}.csv      for each of the 10 ETF tickers
    Group_BMD5302_Robo.xlsx    (the empty template)

Writes:
    Group_BMD5302_Robo_filled.xlsx   (template + NAV data)

The aligned window is 2013-06-01 -> 2026-04-01 (155 monthly rows),
chosen because BNDX is the binding constraint at 155 rows of history.
Rows outside this window are silently dropped per ticker.
"""

import csv
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

TICKERS = ["URTH", "AOA", "XLV", "SPY", "VNQ", "QQQ", "EMB", "BNDX", "AAXJ", "VT"]
ALIGNED_START = date(2013, 6, 1)
ALIGNED_END = date(2026, 4, 1)
EXPECTED_ROWS = 155

REPO_ROOT = Path(__file__).parent
RAW_DIR = REPO_ROOT / "data" / "raw"
TEMPLATE = REPO_ROOT / "Group_BMD5302_Robo.xlsx"
OUTPUT = REPO_ROOT / "Group_BMD5302_Robo_filled.xlsx"


def read_csv(ticker: str) -> dict[date, float]:
    """Read date->NAV map from data/raw/{ticker}.csv, filtered to aligned window."""
    path = RAW_DIR / f"{ticker}.csv"
    if not path.exists():
        raise FileNotFoundError(f"Missing CSV: {path}")

    out: dict[date, float] = {}
    with path.open() as f:
        reader = csv.DictReader(f)
        for row in reader:
            d = date.fromisoformat(row["date"])
            if ALIGNED_START <= d <= ALIGNED_END:
                out[d] = float(row["nav"])
    return out


def main() -> None:
    print(f"Loading template: {TEMPLATE}")
    if not TEMPLATE.exists():
        raise FileNotFoundError(f"Template not found: {TEMPLATE}")

    wb = load_workbook(TEMPLATE)
    ws = wb["NAV_Data"]

    # Read all 10 CSVs
    data: dict[str, dict[date, float]] = {}
    for t in TICKERS:
        nav_map = read_csv(t)
        if len(nav_map) != EXPECTED_ROWS:
            raise ValueError(
                f"{t}: expected {EXPECTED_ROWS} rows in aligned window, "
                f"got {len(nav_map)}. Check data/raw/{t}.csv."
            )
        data[t] = nav_map
        print(f"  {t}: {len(nav_map)} rows in aligned window")

    # Establish the master date list from the first ticker; verify others match
    master_dates = sorted(data[TICKERS[0]].keys())
    for t in TICKERS[1:]:
        ticker_dates = sorted(data[t].keys())
        if ticker_dates != master_dates:
            mismatch_first = next(
                (i for i in range(len(master_dates))
                 if master_dates[i] != ticker_dates[i]),
                len(master_dates),
            )
            raise ValueError(
                f"{t} dates do not match {TICKERS[0]} at index {mismatch_first}: "
                f"{TICKERS[0]}={master_dates[mismatch_first]} "
                f"{t}={ticker_dates[mismatch_first]}"
            )

    # Clear the placeholder marker in A2
    ws["A2"] = master_dates[0].isoformat()

    # Write dates and NAVs
    for i, d in enumerate(master_dates):
        row = 2 + i
        ws.cell(row=row, column=1, value=d.isoformat())
        for j, t in enumerate(TICKERS):
            ws.cell(row=row, column=2 + j, value=data[t][d])

    print(
        f"\nWrote {len(master_dates)} dates x 10 tickers = "
        f"{len(master_dates) * 10} NAV cells"
    )
    print(f"First date: {master_dates[0]}, last date: {master_dates[-1]}")

    wb.save(OUTPUT)
    print(f"\nSaved: {OUTPUT}")
    print("\nNext steps:")
    print("  1. Open the file in Microsoft Excel.")
    print("  2. All formulas (Log_Returns, Cov_Matrix, GMVP) will auto-compute.")
    print("  3. Run Solver on the Optimal sheet (instructions in the sheet itself).")
    print("  4. Run the GenerateFrontier VBA macro on the Frontier sheet.")
    print("  5. Export reconciliation CSVs and drop into data/reconciliation/.")


if __name__ == "__main__":
    main()
