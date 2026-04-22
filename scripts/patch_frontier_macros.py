"""
One-shot utility: overwrite the VBA macro source text embedded on the
Frontier and Frontier_Short sheets of the team's workbook, applying
the two fixes the user specified:

  (1) Reset Optimal!B7:B16 to 0.1 (equal-weight) at the start of every
      iteration inside the For loop, so Solver begins from a neutral
      starting point each row. Prevents the "all rows converge back to
      the tangency" bug we just diagnosed — every row inheriting the
      previous row's Solver result was the root cause.

  (2) Stricter SolverSolve return-code check: accept only codes 0 and 3
      as success. Codes 1 (converged but not optimal), 2 (could not
      converge), and others now flag a convergence_issue in column O.

The macro text is what the user copy-pastes into the VBA editor. This
script does NOT edit the compiled VBA module inside the .xlsm — that
still needs a manual re-paste after the workbook is re-opened in Excel.

Run:
    python scripts/patch_frontier_macros.py

Reads / writes:
    A13_BMD5302_Robo.xlsm   (same file, in-place)

A backup A13_BMD5302_Robo.xlsm.bak was taken before running.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(__file__).resolve().parent.parent / "A13_BMD5302_Robo.xlsm"


# Full replacement text for the Frontier sheet's macro block, starting
# at row 110. Each list item is one row of cell A's text. Empty strings
# produce a blank row (useful for readability).
FRONTIER_MACRO_BLOCK: list[str] = [
    "VBA MACRO TO POPULATE THE FRONTIER  (revised 2026-04)",
    "Open VBA editor (Alt+F11), Insert > Module, paste the code below.",
    "If a previous version is already there, DELETE it and paste this one.",
    "Then run with Alt+F8 > GenerateFrontier > Run.",
    "Requires: Tools > References > check 'Solver'",
    "",
    "Sub GenerateFrontier()",
    "    ' Populate the 100-point efficient frontier on the Frontier sheet.",
    "    ' For each row in B4:B103, target return is in column B; this macro",
    "    ' minimizes portfolio variance subject to that target return,",
    "    ' weights sum to 1, weights in [0, 0.4], and writes weights + stats",
    "    ' back to the same row.",
    "    '",
    "    ' REVISED 2026-04 with two fixes:",
    "    '   (1) Reset Optimal!B7:B16 to 0.1 (equal-weight) at the start of",
    "    '       every iteration so Solver begins from a neutral starting",
    "    '       point each row. Previously the macro inherited the prior",
    "    '       row's weights, and under the 0.4 cap every row converged",
    "    '       back to the same tangency corner (100 identical rows).",
    "    '   (2) Stricter SolverSolve return-code check: only codes 0 and 3",
    "    '       are treated as successful convergence.",
    "",
    "    Dim ws As Worksheet, wsOpt As Worksheet",
    "    Dim i As Long, statusCode As Integer",
    "    Dim weightRange As String",
    "    Dim sumCell As String, varCell As String, retCell As String",
    "    Dim targetReturn As Double",
    "",
    "    Set ws = ThisWorkbook.Sheets(\"Frontier\")",
    "    Set wsOpt = ThisWorkbook.Sheets(\"Optimal\")",
    "",
    "    ' We reuse the Optimal sheet's weight cells (B7:B16) and statistics",
    "    ' as a Solver scratchpad.",
    "",
    "    weightRange = \"Optimal!$B$7:$B$16\"",
    "    sumCell = \"Optimal!$B$18\"",
    "    varCell = \"Optimal!$B$23\"",
    "    retCell = \"Optimal!$B$22\"",
    "",
    "    Application.ScreenUpdating = False",
    "    Application.Calculation = xlCalculationManual",
    "",
    "    For i = 4 To 103",
    "        targetReturn = ws.Cells(i, 2).Value",
    "",
    "        ' FIX (1): reset to equal-weight starting point before each Solver call.",
    "        ' Without this, every row inherits the previous row's result and gets",
    "        ' stuck at the tangency corner under the 0.4 cap.",
    "        wsOpt.Range(\"B7:B16\").Value = 0.1",
    "",
    "        ' Reset Solver",
    "        SolverReset",
    "",
    "        ' Set objective: minimize variance",
    "        SolverOk SetCell:=varCell, MaxMinVal:=2, ByChange:=weightRange, _",
    "                 Engine:=1, EngineDesc:=\"GRG Nonlinear\"",
    "",
    "        ' Constraints",
    "        SolverAdd CellRef:=sumCell, Relation:=2, FormulaText:=\"1\"",
    "        SolverAdd CellRef:=weightRange, Relation:=3, FormulaText:=\"0\"",
    "        SolverAdd CellRef:=weightRange, Relation:=1, FormulaText:=\"0.4\"",
    "        SolverAdd CellRef:=retCell, Relation:=2, FormulaText:=CStr(targetReturn)",
    "",
    "        SolverOptions Convergence:=0.0000001, Iterations:=1000",
    "",
    "        statusCode = SolverSolve(UserFinish:=True)",
    "",
    "        ' Recalc to refresh statistics",
    "        Application.Calculate",
    "",
    "        ' Write results back to Frontier sheet",
    "        ws.Cells(i, 3).Value = wsOpt.Cells(24, 2).Value  ' volatility",
    "        ws.Cells(i, 4).Value = (wsOpt.Cells(22, 2).Value - 0.03) / wsOpt.Cells(24, 2).Value  ' Sharpe",
    "",
    "        Dim k As Integer",
    "        For k = 1 To 10",
    "            ws.Cells(i, 4 + k).Value = wsOpt.Cells(6 + k, 2).Value",
    "        Next k",
    "",
    "        ' FIX (2): stricter convergence check - accept only 0 or 3",
    "        If statusCode <> 0 And statusCode <> 3 Then",
    "            ws.Cells(i, 15).Value = \"convergence_issue:\" & statusCode",
    "        End If",
    "    Next i",
    "",
    "    Application.Calculation = xlCalculationAutomatic",
    "    Application.ScreenUpdating = True",
    "    MsgBox \"Frontier sweep complete. 100 points populated.\"",
    "End Sub",
    "",
    "",
    "Reconciliation export: Save the full B3:N103 range as data/reconciliation/excel_frontier.csv (CSV with headers).",
]


FRONTIER_SHORT_MACRO_BLOCK: list[str] = [
    "VBA MACRO TO POPULATE THE SHORT-ALLOWED FRONTIER  (revised 2026-04)",
    "Open VBA editor (Alt+F11), Insert > Module, paste the code below.",
    "If a previous version is already there, DELETE it and paste this one.",
    "Then run with Alt+F8 > GenerateFrontierShort > Run.",
    "Requires: Tools > References > check 'Solver'",
    "",
    "Sub GenerateFrontierShort()",
    "    ' Populate the 100-point short-allowed efficient frontier.",
    "    ' Same as GenerateFrontier but with bounds [-1, 2] instead of [0, 0.4].",
    "    ' Note: the Optimal sheet's weight cells are reused as Solver scratchpad,",
    "    ' so this macro temporarily overwrites them. Re-run GenerateFrontier",
    "    ' (or re-Solve the Optimal sheet) afterward if you need those weights back.",
    "    '",
    "    ' REVISED 2026-04 with two fixes (identical to GenerateFrontier):",
    "    '   (1) Reset Optimal!B7:B16 to 0.1 at start of every iteration.",
    "    '   (2) Stricter SolverSolve status check: only codes 0 and 3 are success.",
    "",
    "    Dim ws As Worksheet, wsOpt As Worksheet",
    "    Dim i As Long, statusCode As Integer",
    "    Dim weightRange As String",
    "    Dim sumCell As String, varCell As String, retCell As String",
    "    Dim targetReturn As Double",
    "",
    "    Set ws = ThisWorkbook.Sheets(\"Frontier_Short\")",
    "    Set wsOpt = ThisWorkbook.Sheets(\"Optimal\")",
    "",
    "    weightRange = \"Optimal!$B$7:$B$16\"",
    "    sumCell = \"Optimal!$B$18\"",
    "    varCell = \"Optimal!$B$23\"",
    "    retCell = \"Optimal!$B$22\"",
    "",
    "    Application.ScreenUpdating = False",
    "    Application.Calculation = xlCalculationManual",
    "",
    "    For i = 5 To 104",
    "        targetReturn = ws.Cells(i, 2).Value",
    "",
    "        ' FIX (1): reset to equal-weight starting point before each Solver call.",
    "        wsOpt.Range(\"B7:B16\").Value = 0.1",
    "",
    "        SolverReset",
    "",
    "        SolverOk SetCell:=varCell, MaxMinVal:=2, ByChange:=weightRange, _",
    "                 Engine:=1, EngineDesc:=\"GRG Nonlinear\"",
    "",
    "        ' Constraints - note the [-1, 2] bounds vs [0, 0.4] in long-only macro",
    "        SolverAdd CellRef:=sumCell, Relation:=2, FormulaText:=\"1\"",
    "        SolverAdd CellRef:=weightRange, Relation:=3, FormulaText:=\"-1\"",
    "        SolverAdd CellRef:=weightRange, Relation:=1, FormulaText:=\"2\"",
    "        SolverAdd CellRef:=retCell, Relation:=2, FormulaText:=CStr(targetReturn)",
    "",
    "        SolverOptions Convergence:=0.0000001, Iterations:=1000",
    "",
    "        statusCode = SolverSolve(UserFinish:=True)",
    "",
    "        Application.Calculate",
    "",
    "        ws.Cells(i, 3).Value = wsOpt.Cells(24, 2).Value  ' volatility",
    "        ws.Cells(i, 4).Value = (wsOpt.Cells(22, 2).Value - 0.03) / wsOpt.Cells(24, 2).Value",
    "",
    "        Dim k As Integer",
    "        For k = 1 To 10",
    "            ws.Cells(i, 4 + k).Value = wsOpt.Cells(6 + k, 2).Value",
    "        Next k",
    "",
    "        ' FIX (2): stricter convergence check - accept only 0 or 3",
    "        If statusCode <> 0 And statusCode <> 3 Then",
    "            ws.Cells(i, 15).Value = \"convergence_issue:\" & statusCode",
    "        End If",
    "    Next i",
    "",
    "    Application.Calculation = xlCalculationAutomatic",
    "    Application.ScreenUpdating = True",
    "    MsgBox \"Short-allowed frontier sweep complete. 100 points populated.\"",
    "End Sub",
    "",
    "",
    "Reconciliation export: Save the full B4:N104 range as data/reconciliation/excel_frontier_short.csv (CSV with headers).",
]


def _clear_column_a_from_row(ws, first_row: int) -> None:
    """Blank out column A from first_row to the sheet's max_row."""
    for r in range(first_row, ws.max_row + 1):
        ws.cell(row=r, column=1, value=None)


def _write_lines(ws, first_row: int, lines: list[str]) -> None:
    """Write each line of text into column A starting at first_row."""
    for offset, line in enumerate(lines):
        ws.cell(row=first_row + offset, column=1, value=line if line else None)


def main() -> int:
    if not WORKBOOK_PATH.exists():
        print(f"Workbook not found at {WORKBOOK_PATH}", file=sys.stderr)
        return 1

    print(f"Loading {WORKBOOK_PATH.name} (keep_vba=True) ...")
    wb = load_workbook(WORKBOOK_PATH, keep_vba=True)

    for sheet_name, block in (
        ("Frontier", FRONTIER_MACRO_BLOCK),
        ("Frontier_Short", FRONTIER_SHORT_MACRO_BLOCK),
    ):
        ws = wb[sheet_name]
        # Clear old macro text (from row 110 down) then write the new block
        _clear_column_a_from_row(ws, first_row=110)
        _write_lines(ws, first_row=110, lines=block)
        last_row = 110 + len(block) - 1
        print(
            f"  {sheet_name}: wrote {len(block)} lines "
            f"(rows 110 to {last_row})"
        )

    print(f"Saving back to {WORKBOOK_PATH.name} ...")
    wb.save(WORKBOOK_PATH)
    print("Done.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
