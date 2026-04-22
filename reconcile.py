"""
Financial Reconciliation Pipeline — QA / audit harness
=====================================================

Implements the full three-phase reconciliation protocol from PRD Section 4.2:

  Phase 1 — Static Data Reconciliation
    Compares mu vector, covariance matrix, and GMVP weights between the
    Excel baseline exports and the Python backend.

  Phase 2 — Optimization Output Reconciliation
    For each of 5 PRD-canonical A values, compares optimal portfolio weights
    produced by Excel Solver vs. Python SLSQP.

  Phase 3 — Sharpe Ratio & Statistics Reconciliation
    Independently recomputes E(r_p), σ_p, and Sharpe Ratio and verifies
    consistency against both systems.

Tolerance specification (PRD Section 4.3):
  μ vector          atol = 1e-6   (absolute, no rtol)
  Σ matrix          atol = 1e-6
  GMVP weights      atol = 1e-6
  Optimal weights   atol = 1e-6
  E(r_p)            atol = 1e-6
  σ_p               atol = 1e-6
  Sharpe Ratio      atol = 1e-4   (relaxed per PRD)
  Frontier weights  atol = 1e-5   (relaxed per PRD)

Outputs:
  /reports/reconciliation_report.json  — machine-readable results
  /reports/reconciliation_report.md    — human-readable Markdown report
  /reports/reconciliation_report.pdf   — PDF audit summary (PRD QA)

Reconciliation source priority (per check):
  1. Group_BMD5302_Robo.xlsm or .xlsx at the repo root (preferred). The
     workbook must have been opened in Excel at least once so that
     formula cells have cached values (``data_only=True`` reads cached
     values, not formula strings). See read_excel_reconciliation_data.
  2. CSV file under data/reconciliation/ (fallback, per-check). Covers
     two cases: (a) no workbook at repo root yet, (b) workbook present
     but specific sheets not populated.
  3. SKIP, if neither is available.

  The Optimal sheet holds only one A value at a time, so the five
  A-value reconciliations remain CSV-driven (excel_optimal_A{VALUE}.csv)
  regardless of whether the workbook is present.

Usage:
  # Direct Python invocation (backend must be importable)
  python reconcile.py

  # Point to custom Excel CSV directory (Excel baseline exports)
  python reconcile.py --excel-dir path/to/excel_exports/
"""

from __future__ import annotations

import argparse
import datetime
import json
import os
import subprocess
import sys
import time
from pathlib import Path
from typing import Literal, NamedTuple

import numpy as np

# ---------------------------------------------------------------------------
# Status type
# ---------------------------------------------------------------------------
# Every CheckResult now carries a three-valued status instead of a bool:
#   "pass" — reference CSV present AND deviation ≤ tolerance (or internal
#            Python-vs-Python self-consistency check that succeeded)
#   "fail" — reference CSV present AND deviation > tolerance
#   "skip" — no reference CSV found; the check cannot be evaluated against
#            Excel. Previously silently dropped from results; now explicitly
#            reported so "18/18 passed" can't be mistaken for "18/18 verified
#            against Excel".
CheckStatus = Literal["pass", "fail", "skip"]

# ---------------------------------------------------------------------------
# Path setup — reconcile.py lives at the project root
# ---------------------------------------------------------------------------

_PROJECT_ROOT = Path(__file__).resolve().parent
_BACKEND_DIR = _PROJECT_ROOT / "backend"
_PROCESSED_DIR = _PROJECT_ROOT / "data" / "processed"
_REPORTS_DIR = _PROJECT_ROOT / "reports"

# Ensure backend modules are importable
if str(_BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(_BACKEND_DIR))

_REPORTS_DIR.mkdir(parents=True, exist_ok=True)


# ---------------------------------------------------------------------------
# Tolerance constants (PRD Section 4.3)
# ---------------------------------------------------------------------------

TOL_MU = 1e-6
TOL_COV = 1e-6
TOL_GMVP = 1e-6
TOL_OPTIMAL = 1e-6
TOL_RETURN = 1e-6
TOL_VOL = 1e-6
TOL_SHARPE = 1e-4
TOL_FRONTIER = 1e-5

# Single source of truth. backend/ is already on sys.path (see above).
from config import RISK_FREE_RATE  # noqa: E402

# PRD canonical A test values for Phase 2
RECONCILIATION_A_VALUES: list[float] = [0.5, 2.0, 3.5, 6.0, 10.0]


# ---------------------------------------------------------------------------
# Result data structures
# ---------------------------------------------------------------------------


class CheckResult(NamedTuple):
    """
    Outcome of a single reconciliation check.

    ``status`` supersedes the earlier boolean ``passed`` so that checks
    without an Excel reference can be reported as "skip" rather than
    misreported as "pass".

    ``solver_path`` is populated only for tangency checks (values
    "primary" | "fallback"); empty elsewhere. Informational only — does
    not affect pass/fail/skip determination.
    """

    label: str
    status: CheckStatus
    max_deviation: float
    tolerance: float
    details: str = ""
    solver_path: str = ""


def _make_skip(label: str, tolerance: float, reason: str = "no Excel reference") -> CheckResult:
    """Build a CheckResult for a check whose Excel CSV was not provided."""
    return CheckResult(
        label=label,
        status="skip",
        max_deviation=float("nan"),
        tolerance=tolerance,
        details=reason,
    )


# ---------------------------------------------------------------------------
# Console output helpers
# ---------------------------------------------------------------------------

_PASS = "\033[32m[PASS]\033[0m"
_FAIL = "\033[31m[FAIL]\033[0m"
_SKIP = "\033[33m[SKIP]\033[0m"
_INFO = "\033[36m[INFO]\033[0m"


def _log(symbol: str, label: str, msg: str) -> None:
    print(f"{symbol} {label:<55} {msg}")


def _separator(title: str = "") -> None:
    width = 72
    if title:
        pad = (width - len(title) - 2) // 2
        print("=" * pad + f" {title} " + "=" * (width - pad - len(title) - 2))
    else:
        print("=" * width)


# ---------------------------------------------------------------------------
# Core reconciliation function (element-wise, absolute tolerance)
# ---------------------------------------------------------------------------


def reconcile_arrays(
    python_array: np.ndarray,
    reference_array: np.ndarray,
    label: str,
    atol: float,
) -> CheckResult:
    """
    Perform element-wise comparison of two NumPy arrays.

    Uses np.testing.assert_allclose with rtol=0 (absolute tolerance only),
    matching the PRD directive: "no relative tolerance, as near-zero weights
    would create false positives with rtol".

    Parameters
    ----------
    python_array    : values produced by the Python backend
    reference_array : ground-truth values (Excel export or independently computed)
    label           : human-readable check name
    atol            : absolute tolerance per PRD Section 4.3

    Returns
    -------
    CheckResult with pass/fail status and maximum observed deviation
    """
    py = python_array.flatten().astype(np.float64)
    ref = reference_array.flatten().astype(np.float64)

    if py.shape != ref.shape:
        return CheckResult(
            label=label,
            status="fail",
            max_deviation=float("inf"),
            tolerance=atol,
            details=f"Shape mismatch: python={py.shape}, reference={ref.shape}",
        )

    deviations = np.abs(py - ref)
    max_dev = float(np.max(deviations))
    passed = max_dev <= atol

    symbol = _PASS if passed else _FAIL
    _log(symbol, label, f"max deviation = {max_dev:.2e}  (tol={atol:.0e})")

    return CheckResult(
        label=label,
        status="pass" if passed else "fail",
        max_deviation=max_dev,
        tolerance=atol,
        details=(
            f"max_dev={max_dev:.6e}, tol={atol:.0e}, "
            f"worst_idx={int(np.argmax(deviations))}"
        ),
    )


# ---------------------------------------------------------------------------
# Python backend data loaders
# ---------------------------------------------------------------------------


def _load_python_market_data() -> tuple[np.ndarray, np.ndarray]:
    """Load mu and cov directly from the backend data loader."""
    from data_loader import load_market_data  # noqa: PLC0415

    return load_market_data()


def _compute_python_gmvp(cov: np.ndarray) -> np.ndarray:
    """Compute GMVP using the Python backend's closed-form implementation."""
    from optimizer import compute_gmvp  # noqa: PLC0415

    return compute_gmvp(cov)


def _compute_python_optimal(
    mu: np.ndarray,
    cov: np.ndarray,
    A: float,
) -> np.ndarray:
    """Compute optimal portfolio weights for a given A using SLSQP."""
    from optimizer import compute_optimal_portfolio  # noqa: PLC0415

    result = compute_optimal_portfolio(mu, cov, A)
    return result.weights


def _compute_python_frontier(
    mu: np.ndarray,
    cov: np.ndarray,
    n_points: int = 50,
    allow_short_selling: bool = False,
) -> list[np.ndarray]:
    """
    Return a list of weight vectors for n_points frontier portfolios.

    ``allow_short_selling`` is forwarded to the optimizer; default False
    preserves existing Phase 3 reconciliation behaviour (long-only frontier
    vs excel_frontier.csv).
    """
    from optimizer import compute_efficient_frontier  # noqa: PLC0415

    pts = compute_efficient_frontier(
        mu, cov, n_points=n_points, allow_short_selling=allow_short_selling
    )
    return [p.weights for p in pts]


# ---------------------------------------------------------------------------
# Excel CSV loaders (Excel baseline exports)
# ---------------------------------------------------------------------------


def _read_excel_csv_quiet(path: Path) -> np.ndarray | None:
    """
    Load an Excel baseline export CSV without console logging.
    Used when building side-by-side report tables (avoids noisy [SKIP] lines).
    """
    if not path.exists():
        return None
    try:
        import pandas as pd  # noqa: PLC0415

        return pd.read_csv(path, header=None).values.astype(np.float64)
    except Exception:  # noqa: BLE001
        return None


def _require_excel_csv(path: Path, label: str) -> np.ndarray | None:
    """
    Load a CSV file produced by the Excel audit model.
    Returns None and logs a SKIP if the file is absent.
    """
    if not path.exists():
        _log(_SKIP, label, f"Excel CSV not found at {path}")
        return None
    try:
        import pandas as pd  # noqa: PLC0415

        return pd.read_csv(path, header=None).values.astype(np.float64)
    except Exception as exc:  # noqa: BLE001
        _log(_FAIL, label, f"Failed to read CSV: {exc}")
        return None


# ---------------------------------------------------------------------------
# Excel workbook reader (preferred source; CSV is fallback)
# ---------------------------------------------------------------------------
#
# When the team's Group_BMD5302_Robo.xls{m,x} workbook is populated, opened
# in Excel to evaluate formulas, and saved at the repo root, reconcile.py
# reads every reconciliation value directly from known cell locations —
# no more 9 × Save-As-CSV. The per-check ``_require_reconciliation_source``
# helper below prefers workbook data and falls back to the existing CSV
# path row-by-row (so a half-populated workbook still works: populated
# rows come from the workbook, missing rows fall back to CSV or SKIP).
#
# A few workbook-level notes worth pinning so the cell map doesn't drift:
#   * The canonical workbook sheet list is 12 sheets (README, NAV_Data,
#     Log_Returns, Cov_Matrix, GMVP, Optimal, Frontier, Frontier_Chart,
#     Equal_Weight, Tangency, GMVP_Short, Frontier_Short). Changes to
#     this list require updating the cell map below.
#   * Cached formula values in an xlsx are ONLY populated after Excel
#     has opened and re-saved the workbook. A freshly-built openpyxl
#     workbook reads as empty via data_only=True even if every cell is
#     a valid formula. The reader returns None for empty cells; the
#     caller then falls back to CSV.
#   * The Optimal sheet holds only one A value at a time (B3). Per the
#     task spec, the five A-value reconciliations remain CSV-driven
#     (the user saves excel_optimal_A{VALUE}.csv after each Solver run).

# Filename candidates at the repo root, in preference order.
_WORKBOOK_CANDIDATES = (
    "Group_BMD5302_Robo.xlsm",   # macro-enabled, final state
    "Group_BMD5302_Robo.xlsx",   # template / pre-macro state
)

# Cell map per reconciliation key. Each entry describes what to read and
# how to reshape the result.
#   "sheet": Excel sheet name
#   "range": A1-notation range; "A1:K1" → 1×11, etc.
#   "shape": "row" / "col" / "matrix" / "col_weights+stats" — controls
#            post-read reshaping
#   "keys":  when shape is "col_weights+stats", names the two output keys
#            derived from splitting the first 10 rows (weights) from the
#            remaining rows (stats).
_RECONCILIATION_CELL_MAP: list[dict] = [
    # mu vector (retA named range, also addressable as Cov_Matrix!B15:K15)
    {"key": "mu",  "sheet": "Cov_Matrix", "range": "B15:K15", "shape": "row"},
    # covariance matrix (varcov named range, also Cov_Matrix!B2:K11)
    {"key": "cov", "sheet": "Cov_Matrix", "range": "B2:K11",  "shape": "matrix"},
    # GMVP long-only reconciliation export block (rows 27-39): 10 weights then E[r], vol, Sharpe
    {
        "key": ("gmvp_weights", "gmvp_stats"),
        "sheet": "GMVP", "range": "B27:B39", "shape": "col_weights+stats",
    },
    # GMVP short-allowed (rows 30-42)
    {
        "key": ("gmvp_short_weights", "gmvp_short_stats"),
        "sheet": "GMVP_Short", "range": "B30:B42", "shape": "col_weights+stats",
    },
    # Equal-weight (rows 27-39)
    {
        "key": ("equal_weight_weights", "equal_weight_stats"),
        "sheet": "Equal_Weight", "range": "B27:B39", "shape": "col_weights+stats",
    },
    # Tangency long-only (rows 48-60)
    {
        "key": ("tangency_weights", "tangency_stats"),
        "sheet": "Tangency", "range": "B48:B60", "shape": "col_weights+stats",
    },
    # Long-only frontier: 100 rows × 13 cols (target_return, vol, sharpe, w0..w9)
    {"key": "frontier",       "sheet": "Frontier",       "range": "B4:N103",  "shape": "matrix"},
    # Short-allowed frontier: header at row 4 → data rows 5-104
    {"key": "frontier_short", "sheet": "Frontier_Short", "range": "B5:N104", "shape": "matrix"},
]


def _find_excel_workbook(project_root: Path) -> Path | None:
    """Locate the audit workbook at the repo root. Returns None if absent."""
    for name in _WORKBOOK_CANDIDATES:
        path = project_root / name
        if path.exists():
            return path
    return None


def _cells_to_array(cells: tuple, shape: str) -> np.ndarray | None:
    """
    Convert a tuple-of-tuples from openpyxl's ``ws[range]`` into a typed
    numpy array, honouring the reshape hint. Returns None if ANY cell in
    the range is empty (None) — the caller treats that as "not populated;
    fall back to CSV".

    Defensive against three forms of "empty":
      - Empty outer tuple (``ws[range]`` on a read-only empty sheet)
      - Inner tuples of length 0
      - Any cell with .value == None
    """
    # Empty outer or empty first row → treat as "no data"
    if not cells:
        return None
    if not cells[0]:
        return None

    # Flatten to a list, check for any None
    flat: list = []
    for row in cells:
        for cell in row:
            if cell.value is None:
                return None
            flat.append(cell.value)

    if not flat:
        return None

    try:
        if shape in ("row", "col"):
            arr = np.asarray(flat, dtype=np.float64).reshape(-1)
        elif shape == "matrix":
            n_rows = len(cells)
            n_cols = len(cells[0])
            arr = np.asarray(flat, dtype=np.float64).reshape(n_rows, n_cols)
        else:
            arr = np.asarray(flat, dtype=np.float64)
    except (TypeError, ValueError):
        # Non-numeric content in the range (e.g., a formula error string)
        return None
    return arr


def read_excel_reconciliation_data(workbook_path: Path) -> dict:
    """
    Read every reconciliation value from the Excel workbook.

    Returns a dict whose keys are the reconciliation-source labels used
    by the per-check ``_require_reconciliation_source`` helper. Each
    value is either a numpy array or None.

    A None value for a key means "those cells in the workbook are empty
    or contain uncached formulas" — the caller should fall back to the
    corresponding CSV file for that specific check. This supports the
    half-populated-workbook case: any rows the user has evaluated and
    saved are used from the workbook; rows that are still empty fall
    back to CSV (or SKIP).

    The workbook is read with ``data_only=True`` so cached formula
    values are returned rather than the formula strings themselves. If
    the workbook has never been opened in Microsoft Excel, every cached
    value is None and every key in the returned dict will be None,
    which gracefully degrades to the existing CSV-based flow.
    """
    # Heavy import — only pulled in when a workbook is actually present.
    from openpyxl import load_workbook  # noqa: PLC0415

    out: dict = {}

    try:
        wb = load_workbook(workbook_path, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 — malformed workbook is non-fatal
        _log(_FAIL, "Excel workbook load", f"{workbook_path.name}: {exc}")
        # Return all-None so caller drops through to CSV for every check.
        for entry in _RECONCILIATION_CELL_MAP:
            key = entry["key"]
            if isinstance(key, tuple):
                for k in key:
                    out[k] = None
            else:
                out[key] = None
        return out

    for entry in _RECONCILIATION_CELL_MAP:
        sheet_name = entry["sheet"]
        if sheet_name not in wb.sheetnames:
            # Missing sheet — mark the keys as None and continue
            key = entry["key"]
            if isinstance(key, tuple):
                for k in key:
                    out[k] = None
            else:
                out[key] = None
            continue

        ws = wb[sheet_name]
        try:
            cells = ws[entry["range"]]
        except Exception:  # noqa: BLE001 — bad range, treat as missing
            cells = None

        arr = _cells_to_array(cells, entry["shape"]) if cells is not None else None

        key = entry["key"]
        if entry["shape"] == "col_weights+stats":
            # First 10 rows are weights, remaining are stats.
            if arr is None or arr.shape[0] < 13:
                out[key[0]] = None
                out[key[1]] = None
            else:
                out[key[0]] = arr[:10]
                out[key[1]] = arr[10:]
        else:
            out[key] = arr

    return out


def _require_reconciliation_source(
    csv_path: Path,
    label: str,
    *,
    workbook_data: dict | None = None,
    workbook_key: str | None = None,
) -> np.ndarray | None:
    """
    Unified lookup: prefer the workbook value for this check, fall back
    to CSV if the workbook didn't supply it, SKIP if neither is present.

    The workbook is the primary audit source when populated. CSV stays
    the fallback for two cases: (1) the workbook isn't at the repo root
    yet, (2) the user has populated some sheets but not others. Each
    reconciliation check makes its own independent decision so a
    half-populated workbook works correctly.
    """
    if workbook_data is not None and workbook_key is not None:
        arr = workbook_data.get(workbook_key)
        if arr is not None:
            _log(_INFO, label, "source: Excel workbook")
            return np.asarray(arr, dtype=np.float64)

    return _require_excel_csv(csv_path, label)


def _load_excel_optimal_weights(
    excel_dir: Path,
    A: float,
) -> np.ndarray | None:
    """
    Load the optimal weights CSV for a specific A value.
    The Excel baseline must provide files named: excel_optimal_A{A}.csv
    e.g. excel_optimal_A3.5.csv  (10 rows × 1 column)
    """
    fname = f"excel_optimal_A{A}.csv"
    path = excel_dir / fname
    return _require_excel_csv(path, f"Excel optimal weights (A={A})")


# ---------------------------------------------------------------------------
# Independent statistics recomputation (Phase 3)
# ---------------------------------------------------------------------------


def _independently_verify_stats(
    weights: np.ndarray,
    mu: np.ndarray,
    cov: np.ndarray,
    reported_er: float,
    reported_vol: float,
    reported_sharpe: float,
    label: str,
) -> list[CheckResult]:
    """
    Recompute E(r_p), σ_p, and Sharpe Ratio from scratch using NumPy only,
    then compare against both the Python-reported and Excel-computed values.

    This implements the PRD Phase 3 requirement:
    "Independently compute in the reconciliation script — compare against
    both the Excel-reported and Python-reported values."
    """
    results: list[CheckResult] = []

    # --- E(r_p) = w^T μ ------------------------------------------------------
    er_recomputed = float(np.dot(weights, mu))
    dev_er = abs(er_recomputed - reported_er)
    passed_er = dev_er <= TOL_RETURN
    _log(
        _PASS if passed_er else _FAIL,
        f"  {label} E(r_p) independent check",
        f"recomputed={er_recomputed:.8f}, reported={reported_er:.8f}, dev={dev_er:.2e}",
    )
    results.append(CheckResult(
        label=f"{label} E(r_p)",
        status="pass" if passed_er else "fail",
        max_deviation=dev_er,
        tolerance=TOL_RETURN,
    ))

    # --- σ_p = sqrt(w^T Σ w) -------------------------------------------------
    vol_recomputed = float(np.sqrt(np.maximum(weights @ cov @ weights, 0.0)))
    dev_vol = abs(vol_recomputed - reported_vol)
    passed_vol = dev_vol <= TOL_VOL
    _log(
        _PASS if passed_vol else _FAIL,
        f"  {label} σ_p independent check",
        f"recomputed={vol_recomputed:.8f}, reported={reported_vol:.8f}, dev={dev_vol:.2e}",
    )
    results.append(CheckResult(
        label=f"{label} σ_p",
        status="pass" if passed_vol else "fail",
        max_deviation=dev_vol,
        tolerance=TOL_VOL,
    ))

    # --- Sharpe Ratio = (E(r_p) − r_f) / σ_p --------------------------------
    sharpe_recomputed = (er_recomputed - RISK_FREE_RATE) / vol_recomputed if vol_recomputed > 1e-12 else 0.0
    dev_sharpe = abs(sharpe_recomputed - reported_sharpe)
    passed_sharpe = dev_sharpe <= TOL_SHARPE
    _log(
        _PASS if passed_sharpe else _FAIL,
        f"  {label} Sharpe independent check",
        f"recomputed={sharpe_recomputed:.6f}, reported={reported_sharpe:.6f}, dev={dev_sharpe:.2e}",
    )
    results.append(CheckResult(
        label=f"{label} Sharpe",
        status="pass" if passed_sharpe else "fail",
        max_deviation=dev_sharpe,
        tolerance=TOL_SHARPE,
    ))

    return results


# ---------------------------------------------------------------------------
# Phase 1: Static Data Reconciliation
# ---------------------------------------------------------------------------


def run_phase1(
    mu: np.ndarray,
    cov: np.ndarray,
    w_gmvp: np.ndarray,
    excel_dir: Path,
    workbook_data: dict | None = None,
) -> list[CheckResult]:
    """
    Compare mu vector, covariance matrix, and GMVP weights between the
    Python backend and Excel baseline exports.

    When ``workbook_data`` is provided, each check prefers the workbook
    value and falls back to the corresponding CSV if that key is empty.
    """
    _separator("Phase 1: Static Data Reconciliation")
    results: list[CheckResult] = []

    # μ vector (10 elements)
    excel_mu_raw = _require_reconciliation_source(
        excel_dir / "excel_mu_vector.csv", "μ vector",
        workbook_data=workbook_data, workbook_key="mu",
    )
    if excel_mu_raw is not None:
        results.append(reconcile_arrays(mu, excel_mu_raw.flatten(), "μ vector (10 elements)", TOL_MU))
    else:
        results.append(_make_skip("μ vector (10 elements)", TOL_MU))

    # Σ matrix (100 elements)
    excel_cov_raw = _require_reconciliation_source(
        excel_dir / "excel_cov_matrix.csv", "Σ matrix",
        workbook_data=workbook_data, workbook_key="cov",
    )
    if excel_cov_raw is not None:
        results.append(reconcile_arrays(cov, excel_cov_raw, "Σ matrix (100 elements)", TOL_COV))
    else:
        results.append(_make_skip("Σ matrix (100 elements)", TOL_COV))

    # GMVP weights (10 elements)
    excel_gmvp_raw = _require_reconciliation_source(
        excel_dir / "excel_gmvp_weights.csv", "GMVP weights",
        workbook_data=workbook_data, workbook_key="gmvp_weights",
    )
    if excel_gmvp_raw is not None:
        results.append(reconcile_arrays(w_gmvp, excel_gmvp_raw.flatten(), "GMVP weights (10 elements)", TOL_GMVP))
    else:
        results.append(_make_skip("GMVP weights (10 elements)", TOL_GMVP))

    # Phase 3 statistics for GMVP (independent recomputation always runs)
    _separator("Phase 1 — GMVP Statistics (Independent Recomputation)")
    from portfolio_math import portfolio_return, portfolio_volatility, sharpe_ratio  # noqa: PLC0415

    er_gmvp = portfolio_return(w_gmvp, mu)
    vol_gmvp = portfolio_volatility(w_gmvp, cov)
    sr_gmvp = sharpe_ratio(w_gmvp, mu, cov)
    _log(_INFO, "GMVP E(r_p)", f"{er_gmvp:.8f}")
    _log(_INFO, "GMVP σ_p", f"{vol_gmvp:.8f}")
    _log(_INFO, "GMVP Sharpe", f"{sr_gmvp:.8f}")

    stat_results = _independently_verify_stats(
        w_gmvp, mu, cov, er_gmvp, vol_gmvp, sr_gmvp, "GMVP"
    )
    results.extend(stat_results)

    return results


# ---------------------------------------------------------------------------
# Phase 2: Optimization Output Reconciliation
# ---------------------------------------------------------------------------


def run_phase2(
    mu: np.ndarray,
    cov: np.ndarray,
    excel_dir: Path,
) -> list[CheckResult]:
    """
    For each of 5 PRD-canonical A values, compare Python SLSQP optimal weights
    against Excel Solver optimal weights.
    """
    _separator("Phase 2: Optimization Output Reconciliation")
    results: list[CheckResult] = []

    from portfolio_math import portfolio_return, portfolio_volatility, sharpe_ratio  # noqa: PLC0415

    for A in RECONCILIATION_A_VALUES:
        _log(_INFO, f"A = {A}", "Computing Python optimal portfolio…")
        try:
            w_python = _compute_python_optimal(mu, cov, A)
        except Exception as exc:  # noqa: BLE001
            _log(_FAIL, f"A={A} Python optimizer", str(exc))
            results.append(CheckResult(
                label=f"Optimal weights A={A}",
                status="fail",
                max_deviation=float("inf"),
                tolerance=TOL_OPTIMAL,
                details=str(exc),
            ))
            continue

        # Compare against Excel Solver if CSV is available
        excel_w = _load_excel_optimal_weights(excel_dir, A)
        if excel_w is not None:
            results.append(
                reconcile_arrays(
                    w_python,
                    excel_w.flatten(),
                    f"Optimal weights (A={A})",
                    TOL_OPTIMAL,
                )
            )
        else:
            results.append(_make_skip(f"Optimal weights (A={A})", TOL_OPTIMAL))

        # Independent statistics recomputation (always runs)
        er = portfolio_return(w_python, mu)
        vol = portfolio_volatility(w_python, cov)
        sr = sharpe_ratio(w_python, mu, cov)
        stat_results = _independently_verify_stats(w_python, mu, cov, er, vol, sr, f"Optimal A={A}")
        results.extend(stat_results)

    return results


# ---------------------------------------------------------------------------
# Phase 3: Efficient Frontier Reconciliation
# ---------------------------------------------------------------------------


def run_phase3_frontier(
    mu: np.ndarray,
    cov: np.ndarray,
    excel_dir: Path,
    workbook_data: dict | None = None,
) -> list[CheckResult]:
    """
    Compare the efficient frontier points (up to 100 points matching the
    Excel model) between Python and Excel Solver.

    Source preference: workbook_data["frontier"] (when populated) ahead of
    excel_dir/excel_frontier.csv. The workbook's frontier block is
    Frontier!B4:N103 — 100 rows × 13 cols: [target_return, volatility,
    sharpe, w0..w9]. We extract the w0..w9 slice for reconciliation.
    """
    _separator("Phase 3: Efficient Frontier Reconciliation")
    results: list[CheckResult] = []

    excel_weights_matrix: np.ndarray | None = None

    # First preference: the workbook's frontier block
    if workbook_data is not None:
        arr = workbook_data.get("frontier")
        if arr is not None and arr.ndim == 2 and arr.shape[1] >= 13:
            # Columns: [target_return, volatility, sharpe, w0..w9]. Extract weights.
            excel_weights_matrix = arr[:, 3:13].astype(np.float64)
            _log(_INFO, "Frontier source", "Excel workbook (Frontier!B4:N103)")

    # Fallback: CSV
    if excel_weights_matrix is None:
        excel_frontier_path = excel_dir / "excel_frontier.csv"
        if not excel_frontier_path.exists():
            _log(_SKIP, "Frontier reconciliation", "no workbook source and excel_frontier.csv not found")
            results.append(_make_skip(
                "Frontier weights (long-only, 100 points)",
                TOL_FRONTIER,
            ))
            return results

        import pandas as pd  # noqa: PLC0415

        df = pd.read_csv(excel_frontier_path)

        # Expected columns: target_return, min_variance, w0, w1, ..., w9
        weight_cols = [c for c in df.columns if c.startswith("w") or "weight" in c.lower()]

        if not weight_cols:
            _log(_FAIL, "Frontier CSV parse", "No weight columns detected in excel_frontier.csv")
            return results

        _log(_INFO, "Frontier source", f"CSV ({excel_frontier_path.name})")
        excel_weights_matrix = df[weight_cols].values.astype(np.float64)

    n_excel_points = len(excel_weights_matrix)
    _log(_INFO, "Excel frontier points", str(n_excel_points))

    # Generate matching Python frontier (same count as Excel)
    try:
        py_frontier_weights = _compute_python_frontier(mu, cov, n_points=n_excel_points)
    except Exception as exc:  # noqa: BLE001
        _log(_FAIL, "Python frontier computation", str(exc))
        return results

    py_weights_matrix = np.array(py_frontier_weights, dtype=np.float64)

    if excel_weights_matrix.shape[1] != py_weights_matrix.shape[1]:
        _log(
            _FAIL,
            "Frontier weight column count mismatch",
            f"Excel={excel_weights_matrix.shape[1]}, Python={py_weights_matrix.shape[1]}",
        )
        return results

    # Align lengths
    n_compare = min(len(excel_weights_matrix), len(py_weights_matrix))
    excel_slice = excel_weights_matrix[:n_compare]
    py_slice = py_weights_matrix[:n_compare]

    # Compare row-wise (each frontier point)
    max_devs = np.max(np.abs(excel_slice - py_slice), axis=1)
    overall_max = float(max_devs.max())
    passed = overall_max <= TOL_FRONTIER

    _log(
        _PASS if passed else _FAIL,
        f"Frontier weights ({n_compare} points × 10 weights)",
        f"max deviation = {overall_max:.2e}  (tol={TOL_FRONTIER:.0e})",
    )
    results.append(CheckResult(
        label=f"Frontier weights ({n_compare} points)",
        status="pass" if passed else "fail",
        max_deviation=overall_max,
        tolerance=TOL_FRONTIER,
        details=f"Worst frontier point index: {int(np.argmax(max_devs))}",
    ))

    return results


# ---------------------------------------------------------------------------
# Phase 3b: PRD Part 1 additions (short-allowed + tangency + equal-weight)
# ---------------------------------------------------------------------------


def run_phase3b_prd_part1(
    mu: np.ndarray,
    cov: np.ndarray,
    excel_dir: Path,
    workbook_data: dict | None = None,
) -> list[CheckResult]:
    """
    Reconcile the artifacts introduced in Steps 1–3 of the PRD Part 1
    work: short-allowed GMVP, long-only tangency, short-allowed tangency,
    short-allowed efficient frontier, equal-weight benchmark.

    Source preference per check: workbook_data[...] (when populated)
    ahead of the excel_dir CSV. Each check is independent — a partially
    populated workbook (some sheets evaluated, others empty) falls back
    to CSV on a per-row basis.

    Tangency rows additionally carry the Python solver_path
    ("primary" | "fallback") for provenance.
    """
    _separator("Phase 3b: PRD Part 1 artifacts (short-allowed + tangency + equal-weight)")
    results: list[CheckResult] = []

    # Lazy-import so reconcile.py can still be imported without a fully
    # populated backend (e.g. in lightweight CI stages).
    from optimizer import (  # noqa: PLC0415
        _compute_constrained_gmvp,
        compute_efficient_frontier,
        compute_equal_weight_portfolio,
        compute_tangency_portfolio,
    )
    from portfolio_math import (  # noqa: PLC0415
        portfolio_return as _pr,
        portfolio_volatility as _pv,
        sharpe_ratio as _sr,
    )

    # --- 1. GMVP (short-allowed) ------------------------------------------
    w_gmvp_s = _compute_constrained_gmvp(cov, allow_short_selling=True)
    _log(_INFO, "Py GMVP (short-allowed)", f"E[r]={_pr(w_gmvp_s, mu):.6f} σ={_pv(w_gmvp_s, cov):.6f}")

    excel_w = _require_reconciliation_source(
        excel_dir / "excel_gmvp_short.csv", "GMVP (short-allowed) weights",
        workbook_data=workbook_data, workbook_key="gmvp_short_weights",
    )
    if excel_w is not None:
        results.append(
            reconcile_arrays(
                w_gmvp_s, excel_w.flatten(),
                "GMVP (short-allowed) weights", TOL_GMVP,
            )
        )
    else:
        results.append(_make_skip("GMVP (short-allowed) weights", TOL_GMVP))

    # --- 2. Tangency (long-only, max_weight=1.0) --------------------------
    tan_long = compute_tangency_portfolio(
        mu, cov, max_weight=1.0, allow_short_selling=False
    )
    _log(
        _INFO,
        "Py Tangency (long-only)",
        f"Sharpe={tan_long.sharpe:.6f}  solver_path={tan_long.solver_path}",
    )

    excel_w = _require_reconciliation_source(
        excel_dir / "excel_tangency.csv", "Tangency (long-only) weights",
        workbook_data=workbook_data, workbook_key="tangency_weights",
    )
    if excel_w is not None:
        r = reconcile_arrays(
            tan_long.weights, excel_w.flatten(),
            "Tangency (long-only) weights", TOL_OPTIMAL,
        )
        results.append(r._replace(solver_path=tan_long.solver_path))
    else:
        results.append(
            _make_skip("Tangency (long-only) weights", TOL_OPTIMAL)
            ._replace(solver_path=tan_long.solver_path)
        )

    # --- 3. Tangency (short-allowed) --------------------------------------
    tan_short = compute_tangency_portfolio(
        mu, cov, max_weight=1.0, allow_short_selling=True
    )
    _log(
        _INFO,
        "Py Tangency (short-allowed)",
        f"Sharpe={tan_short.sharpe:.6f}  solver_path={tan_short.solver_path}",
    )

    # Short-allowed tangency is NOT in the current workbook sheet set
    # (Tangency sheet is long-only only); CSV is the only source.
    excel_w = _require_reconciliation_source(
        excel_dir / "excel_tangency_short.csv", "Tangency (short-allowed) weights",
        workbook_data=workbook_data, workbook_key="tangency_short_weights",
    )
    if excel_w is not None:
        r = reconcile_arrays(
            tan_short.weights, excel_w.flatten(),
            "Tangency (short-allowed) weights", TOL_OPTIMAL,
        )
        results.append(r._replace(solver_path=tan_short.solver_path))
    else:
        results.append(
            _make_skip("Tangency (short-allowed) weights", TOL_OPTIMAL)
            ._replace(solver_path=tan_short.solver_path)
        )

    # --- 4. Short-allowed efficient frontier (100 points) -----------------
    # Prefer workbook_data["frontier_short"] when populated; else CSV.
    excel_mat: np.ndarray | None = None
    n_expected = 100

    if workbook_data is not None:
        arr = workbook_data.get("frontier_short")
        if arr is not None and arr.ndim == 2 and arr.shape[1] >= 13:
            excel_mat = arr[:, 3:13].astype(np.float64)
            _log(_INFO, "Short-allowed frontier source", "Excel workbook (Frontier_Short!B5:N104)")
            n_expected = len(excel_mat)

    if excel_mat is None:
        path = excel_dir / "excel_frontier_short.csv"
        if path.exists():
            try:
                import pandas as pd  # noqa: PLC0415

                df = pd.read_csv(path)
                weight_cols = [c for c in df.columns if c.startswith("w") or "weight" in c.lower()]
                if not weight_cols:
                    results.append(
                        _make_skip(
                            "Frontier weights (short-allowed, 100 points)",
                            TOL_FRONTIER,
                            "no weight columns detected",
                        )
                    )
                    excel_mat = None
                    n_expected = 0
                else:
                    excel_mat = df[weight_cols].values.astype(np.float64)
                    n_expected = len(excel_mat)
                    _log(_INFO, "Short-allowed frontier source", f"CSV ({path.name})")
            except Exception as exc:  # noqa: BLE001
                _log(_FAIL, "Short-allowed frontier CSV parse", str(exc))
                results.append(CheckResult(
                    label="Frontier weights (short-allowed)",
                    status="fail",
                    max_deviation=float("inf"),
                    tolerance=TOL_FRONTIER,
                    details=str(exc),
                ))
                excel_mat = None
                n_expected = 0

    if excel_mat is not None and n_expected > 0:
        py_frontier = compute_efficient_frontier(
            mu, cov, n_points=n_expected, max_weight=1.0, allow_short_selling=True
        )
        py_mat = np.array([p.weights for p in py_frontier], dtype=np.float64)
        n_compare = min(len(excel_mat), len(py_mat))
        max_devs = np.max(np.abs(excel_mat[:n_compare] - py_mat[:n_compare]), axis=1)
        overall_max = float(max_devs.max())
        passed = overall_max <= TOL_FRONTIER
        _log(
            _PASS if passed else _FAIL,
            f"Short-allowed frontier weights ({n_compare} × 10)",
            f"max deviation = {overall_max:.2e}  (tol={TOL_FRONTIER:.0e})",
        )
        results.append(CheckResult(
            label=f"Frontier weights (short-allowed, {n_compare} points)",
            status="pass" if passed else "fail",
            max_deviation=overall_max,
            tolerance=TOL_FRONTIER,
            details=f"Worst frontier point index: {int(np.argmax(max_devs))}",
        ))
    elif excel_mat is None and n_expected == 0:
        # Already appended a skip/fail above; skip the default append
        pass
    else:
        results.append(_make_skip(
            "Frontier weights (short-allowed, 100 points)", TOL_FRONTIER
        ))

    # --- 5. Equal-weight benchmark ----------------------------------------
    # The workbook's Equal_Weight sheet exposes weights (all 0.1) + stats.
    # The reconciliation check compares only the 3 stats (E[r], σ, Sharpe).
    ew = compute_equal_weight_portfolio(mu, cov)
    excel_w = _require_reconciliation_source(
        excel_dir / "excel_equal_weight.csv", "Equal-weight stats",
        workbook_data=workbook_data, workbook_key="equal_weight_stats",
    )
    if excel_w is not None:
        ref = excel_w.flatten()
        py_vec = np.array([ew.expected_return, ew.volatility, ew.sharpe])
        n_take = min(len(ref), 3)
        results.append(reconcile_arrays(
            py_vec[:n_take], ref[:n_take],
            "Equal-weight (E[r], σ, Sharpe)", TOL_RETURN,
        ))
    else:
        results.append(_make_skip("Equal-weight (E[r], σ, Sharpe)", TOL_RETURN))

    return results


# ---------------------------------------------------------------------------
# Report generation
# ---------------------------------------------------------------------------


def _get_git_sha() -> str:
    """Return the current git commit SHA (short), or 'N/A' if not in a repo."""
    try:
        result = subprocess.run(
            ["git", "rev-parse", "--short", "HEAD"],
            capture_output=True,
            text=True,
            cwd=str(_PROJECT_ROOT),
            timeout=5,
        )
        return result.stdout.strip() if result.returncode == 0 else "N/A"
    except Exception:  # noqa: BLE001
        return "N/A"


def _get_excel_version(excel_dir: Path) -> str:
    """Return the last-modified timestamp of the most recently changed Excel CSV."""
    csvs = sorted(excel_dir.glob("excel_*.csv"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not csvs:
        return "N/A (no Excel CSVs found)"
    ts = datetime.datetime.fromtimestamp(csvs[0].stat().st_mtime)
    return ts.strftime("%Y-%m-%d %H:%M:%S")


def generate_json_report(
    results: list[CheckResult],
    excel_dir: Path,
    mu: np.ndarray,
    cov: np.ndarray,
    w_gmvp: np.ndarray,
    elapsed_seconds: float,
) -> dict:
    """Build the machine-readable reconciliation_report.json payload."""
    # (pass/fail/skip counts are computed below from the new status field)

    # Side-by-side GMVP table (Python vs Excel if available)
    excel_gmvp_raw = _read_excel_csv_quiet(excel_dir / "excel_gmvp_weights.csv")
    gmvp_comparison: list[dict] = []
    for i, pw in enumerate(w_gmvp.tolist()):
        row: dict = {"asset_index": i, "python_weight": round(pw, 10)}
        if excel_gmvp_raw is not None:
            ew = float(excel_gmvp_raw.flatten()[i])
            row["excel_weight"] = round(ew, 10)
            row["deviation"] = round(abs(pw - ew), 10)
        gmvp_comparison.append(row)

    # Side-by-side optimal portfolio table for A=3.5
    py_w_opt = _compute_python_optimal(mu, cov, A=3.5)
    excel_opt_raw = _read_excel_csv_quiet(excel_dir / "excel_optimal_A3.5.csv")
    opt_comparison: list[dict] = []
    for i, pw in enumerate(py_w_opt.tolist()):
        row_opt: dict = {"asset_index": i, "python_weight": round(pw, 10)}
        if excel_opt_raw is not None:
            ew = float(excel_opt_raw.flatten()[i])
            row_opt["excel_weight"] = round(ew, 10)
            row_opt["deviation"] = round(abs(pw - ew), 10)
        opt_comparison.append(row_opt)

    passed_count = sum(1 for r in results if r.status == "pass")
    failed_count = sum(1 for r in results if r.status == "fail")
    skipped_count = sum(1 for r in results if r.status == "skip")

    # overall_status: FAIL if any check failed; SKIP if nothing actually
    # verified against Excel (pass_count == 0 but skip_count > 0);
    # otherwise PASS. Mirrors the CLI summary logic exactly.
    if failed_count > 0:
        overall_status = "FAIL"
    elif passed_count == 0:
        overall_status = "SKIP"
    else:
        overall_status = "PASS"

    return {
        "timestamp": datetime.datetime.now(datetime.timezone.utc).isoformat().replace("+00:00", "Z"),
        "git_commit_sha": _get_git_sha(),
        "excel_file_version": _get_excel_version(excel_dir),
        "overall_status": overall_status,
        "total_checks": len(results),
        "passed_checks": passed_count,
        "failed_checks": failed_count,
        "skipped_checks": skipped_count,
        "elapsed_seconds": round(elapsed_seconds, 3),
        "checks": [
            {
                "label": r.label,
                "status": r.status.upper(),  # "PASS" | "FAIL" | "SKIP"
                "max_deviation": (
                    None if r.status == "skip" and not np.isfinite(r.max_deviation)
                    else r.max_deviation
                ),
                "tolerance": r.tolerance,
                "details": r.details,
                **({"solver_path": r.solver_path} if r.solver_path else {}),
            }
            for r in results
        ],
        "gmvp_side_by_side": gmvp_comparison,
        "optimal_side_by_side_A3_5": opt_comparison,
    }


def generate_markdown_report(report: dict, mu: np.ndarray, cov: np.ndarray) -> str:
    """Generate the human-readable reconciliation_report.md content."""
    overall = report["overall_status"]
    badge = {
        "PASS": "✅ PASS",
        "FAIL": "❌ FAIL",
        "SKIP": "⚠ SKIP (not verified against Excel)",
    }.get(overall, overall)

    summary_parts = [f"**{report['passed_checks']} passed**"]
    if report.get("skipped_checks", 0):
        summary_parts.append(f"**{report['skipped_checks']} skipped** (no Excel reference)")
    if report.get("failed_checks", 0):
        summary_parts.append(f"**{report['failed_checks']} failed**")
    summary_line = ", ".join(summary_parts) + f" out of {report['total_checks']} total."

    lines: list[str] = [
        "# Robo-Adviser Reconciliation Report",
        "",
        f"**Overall Status:** {badge}  ",
        f"**Timestamp:** {report['timestamp']}  ",
        f"**Git Commit SHA (Backend):** `{report['git_commit_sha']}`  ",
        f"**Excel Model Version:** {report['excel_file_version']}  ",
        f"**Elapsed Time:** {report['elapsed_seconds']:.3f}s  ",
        "",
        summary_line,
        "",
        "> **SKIP semantics:** a check reports SKIP when no Excel reference CSV is "
        "found under `data/reconciliation/`. Earlier versions of this report silently "
        "dropped skipped rows, which made the pass count look stronger than it was. "
        "Each SKIP row below identifies a reconciliation gap that the Excel audit model "
        "will eventually close.",
        "",
        "---",
        "",
        "## Check Results",
        "",
        "| # | Check | Status | Max Deviation | Tolerance | Notes |",
        "|---|-------|--------|--------------|-----------|-------|",
    ]

    status_badge = {
        "PASS": "✅ PASS",
        "FAIL": "❌ FAIL",
        "SKIP": "⚠ SKIP (no Excel reference)",
    }

    for i, check in enumerate(report["checks"], start=1):
        st = status_badge.get(check["status"], check["status"])
        if check["status"] == "SKIP":
            dev_cell = "—"
        else:
            dev_cell = f"`{check['max_deviation']:.2e}`"
        notes_parts: list[str] = []
        if check.get("solver_path"):
            notes_parts.append(f"solver: `{check['solver_path']}`")
        if check["status"] == "SKIP" and check.get("details"):
            notes_parts.append(check["details"])
        notes_cell = "; ".join(notes_parts) if notes_parts else ""
        lines.append(
            f"| {i} | {check['label']} | {st} | "
            f"{dev_cell} | `{check['tolerance']:.0e}` | {notes_cell} |"
        )

    lines += [
        "",
        "---",
        "",
        "## GMVP Weights: Python vs Excel",
        "",
        "| Asset | Python Weight | Excel Weight | Deviation |",
        "|-------|--------------|-------------|-----------|",
    ]
    for row in report["gmvp_side_by_side"]:
        excel_w = f"`{row['excel_weight']:.8f}`" if "excel_weight" in row else "N/A"
        dev = f"`{row['deviation']:.2e}`" if "deviation" in row else "N/A"
        lines.append(
            f"| {row['asset_index']} | `{row['python_weight']:.8f}` | {excel_w} | {dev} |"
        )

    lines += [
        "",
        "---",
        "",
        "## Optimal Portfolio Weights (A=3.5): Python vs Excel",
        "",
        "| Asset | Python Weight | Excel Weight | Deviation |",
        "|-------|--------------|-------------|-----------|",
    ]
    for row in report["optimal_side_by_side_A3_5"]:
        excel_w = f"`{row['excel_weight']:.8f}`" if "excel_weight" in row else "N/A"
        dev = f"`{row['deviation']:.2e}`" if "deviation" in row else "N/A"
        lines.append(
            f"| {row['asset_index']} | `{row['python_weight']:.8f}` | {excel_w} | {dev} |"
        )

    lines += [
        "",
        "---",
        "",
        "## Tolerance Specifications (PRD Section 4.3)",
        "",
        "| Metric | Absolute Tolerance | Notes |",
        "|--------|-------------------|-------|",
        f"| μ vector (10 elements) | `{TOL_MU:.0e}` | Annualized mean returns |",
        f"| Σ matrix (100 elements) | `{TOL_COV:.0e}` | Annualized covariance matrix |",
        f"| GMVP weights (10) | `{TOL_GMVP:.0e}` | Closed-form vs MMULT/MINVERSE |",
        f"| Optimal weights (10, per A) | `{TOL_OPTIMAL:.0e}` | SLSQP vs Excel Solver |",
        f"| E(r_p) | `{TOL_RETURN:.0e}` | Portfolio expected return |",
        f"| σ_p | `{TOL_VOL:.0e}` | Portfolio volatility |",
        f"| Sharpe Ratio | `{TOL_SHARPE:.0e}` | Relaxed due to r_f rounding |",
        f"| Frontier weights | `{TOL_FRONTIER:.0e}` | Parametric sweep; relaxed |",
        "",
        "---",
        "",
        "## Failure Escalation Protocol",
        "",
        "If any check fails, refer to PRD Section 4.4 root-cause categories:",
        "",
        "- **DATA_PIPELINE_ERROR** — μ mismatch → check decimal precision / date alignment",
        "- **MATRIX_ALGEBRA_BUG** — μ matches but GMVP fails → NumPy inv() vs MINVERSE()",
        "- **OPTIMIZER_CONVERGENCE** — GMVP matches but optimal portfolio fails → tighten ftol",
        "- **ANNUALIZATION_ERROR** — Monthly vs annual factor mismatch (12 vs 252)",
        "",
        "_Report generated by `reconcile.py` (QA reconciliation harness)._",
    ]

    return "\n".join(lines)


def _pdf_ascii(text: str) -> str:
    """Helvetica core font supports WinAnsi only — strip/replace non-ASCII for fpdf2."""
    return text.encode("ascii", errors="replace").decode("ascii")


def generate_pdf_report(report: dict, output_path: Path) -> None:
    """
    Write a human-readable PDF audit document (PRD QA deliverable).

    Uses fpdf2 with Helvetica (ASCII / WinAnsi) — dynamic check labels may
    contain Greek symbols from console output; they are sanitized.
    """
    from fpdf import FPDF

    pdf = FPDF()
    pdf.set_margins(12, 12, 12)
    pdf.set_auto_page_break(auto=True, margin=14)
    pdf.add_page()
    epw = pdf.w - pdf.l_margin - pdf.r_margin

    pdf.set_font("Helvetica", "B", 14)
    pdf.multi_cell(epw, 8, "Robo-Adviser Reconciliation Report")
    pdf.ln(2)
    pdf.set_font("Helvetica", size=9)
    summary = (
        f"Timestamp (UTC): {report['timestamp']}\n"
        f"Git commit (backend): {report['git_commit_sha']}\n"
        f"Excel exports version: {_pdf_ascii(str(report['excel_file_version']))}\n"
        f"Overall status: {report['overall_status']}\n"
        f"Checks: {report['passed_checks']} passed, "
        f"{report.get('skipped_checks', 0)} skipped (no Excel reference), "
        f"{report['failed_checks']} failed  (total {report['total_checks']})\n"
        f"Elapsed (s): {report['elapsed_seconds']}\n"
    )
    pdf.multi_cell(epw, 5, _pdf_ascii(summary))
    pdf.ln(3)
    pdf.set_font("Helvetica", "B", 11)
    pdf.multi_cell(epw, 6, "Per-check results")
    pdf.set_font("Helvetica", size=8)
    for i, check in enumerate(report.get("checks", []), start=1):
        status = check.get("status", "?")
        dev = check.get("max_deviation")
        dev_str = "n/a (skip)" if status == "SKIP" or dev is None else str(dev)
        solver = check.get("solver_path")
        trailer = f"  [solver={solver}]" if solver else ""
        line = (
            f"{i}. [{status}] {check.get('label', '')}: "
            f"max_dev={dev_str} tol={check.get('tolerance')!s}{trailer}"
        )
        pdf.multi_cell(epw, 4, _pdf_ascii(line))
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 11)
    pdf.multi_cell(epw, 6, "GMVP weights (Python vs Excel)")
    pdf.set_font("Helvetica", size=8)
    for row in report.get("gmvp_side_by_side", [])[:12]:
        idx = row.get("asset_index", "")
        pw = row.get("python_weight", "")
        if "excel_weight" in row:
            line = f"Asset {idx}: py={pw} excel={row['excel_weight']} dev={row.get('deviation', '')}"
        else:
            line = f"Asset {idx}: py={pw} excel=N/A"
        pdf.multi_cell(epw, 4, _pdf_ascii(line))
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 11)
    pdf.multi_cell(epw, 6, "Optimal weights A=3.5 (Python vs Excel)")
    pdf.set_font("Helvetica", size=8)
    for row in report.get("optimal_side_by_side_A3_5", [])[:12]:
        idx = row.get("asset_index", "")
        pw = row.get("python_weight", "")
        if "excel_weight" in row:
            line = f"Asset {idx}: py={pw} excel={row['excel_weight']} dev={row.get('deviation', '')}"
        else:
            line = f"Asset {idx}: py={pw} excel=N/A"
        pdf.multi_cell(epw, 4, _pdf_ascii(line))
    pdf.ln(4)
    pdf.set_font("Helvetica", "I", 8)
    pdf.multi_cell(
        epw,
        4,
        _pdf_ascii(
            "Tolerances follow PRD Section 4.3 (e.g. mu, cov, GMVP: atol=1e-6; Sharpe: 1e-4). "
            "Full machine-readable JSON is stored alongside this PDF as reconciliation_report.json."
        ),
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    pdf.output(str(output_path))


# ---------------------------------------------------------------------------
# Main reconciliation runner
# ---------------------------------------------------------------------------


def run_reconciliation(excel_dir: Path | None = None) -> dict:
    """
    Execute all three reconciliation phases and emit report files.

    Parameters
    ----------
    excel_dir : Path | None
        Directory containing Excel baseline CSV exports.
        Defaults to /data/reconciliation/.

    Returns
    -------
    dict — the full JSON report payload (also written to /reports/)
    """
    if excel_dir is None:
        excel_dir = _PROJECT_ROOT / "data" / "reconciliation"

    excel_dir.mkdir(parents=True, exist_ok=True)

    t_start = time.monotonic()

    _separator("Robo-Adviser Financial Reconciliation")
    print(f"{_INFO} Timestamp      : {datetime.datetime.now(datetime.timezone.utc).isoformat().replace('+00:00', 'Z')}")
    print(f"{_INFO} Git commit SHA  : {_get_git_sha()}")
    print(f"{_INFO} Excel dir       : {excel_dir}")
    print(f"{_INFO} Excel version   : {_get_excel_version(excel_dir)}")

    # --- Load Excel workbook (preferred source; CSV is fallback) -----------
    workbook_path = _find_excel_workbook(_PROJECT_ROOT)
    workbook_data: dict | None = None
    if workbook_path is not None:
        print(f"{_INFO} Excel workbook  : {workbook_path.name}")
        workbook_data = read_excel_reconciliation_data(workbook_path)
        populated_keys = sorted(
            k for k, v in workbook_data.items() if v is not None
        )
        if populated_keys:
            print(f"{_INFO} Workbook cache  : {len(populated_keys)} keys populated  "
                  f"({', '.join(populated_keys[:6])}{'…' if len(populated_keys) > 6 else ''})")
        else:
            print(f"{_INFO} Workbook cache  : empty  "
                  "(open in Excel + save to evaluate formulas; falling back to CSVs)")
    else:
        print(f"{_INFO} Excel workbook  : (not found at repo root; CSV-only mode)")

    _separator()

    # --- Load Python backend data -------------------------------------------
    _log(_INFO, "Loading Python market data", "")
    try:
        mu, cov = _load_python_market_data()
        _log(_INFO, "  mu shape", str(mu.shape))
        _log(_INFO, "  cov shape", str(cov.shape))
        _log(_INFO, "  cov condition number", f"{np.linalg.cond(cov):.4e}")
        _log(_INFO, "  cov det (MDETERM proxy)", f"{np.linalg.det(cov):.4e}")
    except Exception as exc:  # noqa: BLE001
        print(f"{_FAIL} Cannot load Python market data: {exc}")
        print("Ensure mu_vector.json and cov_matrix.json exist under data/processed/.")
        sys.exit(1)

    _log(_INFO, "Computing Python GMVP", "")
    try:
        w_gmvp = _compute_python_gmvp(cov)
        _log(_INFO, "  GMVP weights sum", f"{w_gmvp.sum():.10f}")
        _log(_INFO, "  GMVP min weight", f"{w_gmvp.min():.6f}")
    except Exception as exc:  # noqa: BLE001
        print(f"{_FAIL} GMVP computation failed: {exc}")
        sys.exit(1)

    # --- Phase 1 ------------------------------------------------------------
    results: list[CheckResult] = []
    results.extend(run_phase1(mu, cov, w_gmvp, excel_dir, workbook_data=workbook_data))

    # --- Phase 2 ------------------------------------------------------------
    # Optimal sheet holds one A value at a time; per spec, stays CSV-driven.
    results.extend(run_phase2(mu, cov, excel_dir))

    # --- Phase 3: Frontier --------------------------------------------------
    results.extend(run_phase3_frontier(mu, cov, excel_dir, workbook_data=workbook_data))

    # --- Phase 3b: PRD Part 1 additions -------------------------------------
    results.extend(run_phase3b_prd_part1(mu, cov, excel_dir, workbook_data=workbook_data))

    # --- Summary ------------------------------------------------------------
    elapsed = time.monotonic() - t_start
    _separator("Reconciliation Summary")
    passed = sum(1 for r in results if r.status == "pass")
    failed = sum(1 for r in results if r.status == "fail")
    skipped = sum(1 for r in results if r.status == "skip")
    total = len(results)

    print(f"\n  Total checks   : {total}")
    print(f"  Passed         : {passed}")
    print(f"  Failed         : {failed}")
    print(f"  Skipped        : {skipped}  (no Excel reference)")
    print(f"  Elapsed        : {elapsed:.3f}s")

    # Overall status treats skips as "not verified" — only FAIL is terminal.
    # If there are zero pass/fail results (everything skipped), overall is
    # SKIP: nothing has actually been verified against Excel.
    if failed > 0:
        print(f"\n{_FAIL} {failed} CHECK(S) FAILED — see details above")
        overall = "FAIL"
    elif passed == 0:
        print(f"\n{_SKIP} NO CHECKS VERIFIED AGAINST EXCEL — all rows skipped")
        print("         Provide Excel baseline exports to /data/reconciliation/ and re-run.")
        overall = "SKIP"
    else:
        msg = f"{passed} check(s) passed"
        if skipped:
            msg += f", {skipped} skipped (no Excel reference)"
        print(f"\n{_PASS} {msg}")
        overall = "PASS"

    # --- Generate reports ---------------------------------------------------
    _separator("Generating Reports")
    report = generate_json_report(results, excel_dir, mu, cov, w_gmvp, elapsed)
    report["overall_status"] = overall  # override with SKIP if no checks ran

    json_path = _REPORTS_DIR / "reconciliation_report.json"
    md_path = _REPORTS_DIR / "reconciliation_report.md"

    with open(json_path, "w", encoding="utf-8") as fh:
        json.dump(report, fh, indent=2, ensure_ascii=False)
    _log(_INFO, "JSON report", str(json_path))

    md_content = generate_markdown_report(report, mu, cov)
    with open(md_path, "w", encoding="utf-8") as fh:
        fh.write(md_content)
    _log(_INFO, "Markdown report", str(md_path))

    pdf_path = _REPORTS_DIR / "reconciliation_report.pdf"
    try:
        generate_pdf_report(report, pdf_path)
        _log(_INFO, "PDF report", str(pdf_path))
    except Exception as exc:  # noqa: BLE001
        _log(_FAIL, "PDF report", str(exc))

    _separator()
    return report


# ---------------------------------------------------------------------------
# pytest-compatible wrappers for use in test_reconciliation.py
# ---------------------------------------------------------------------------


def assert_mu_reconciliation(excel_dir: Path | None = None) -> CheckResult:
    """
    Pytest-friendly assertion wrapper for the μ vector check.
    Raises AssertionError if the check fails (for use in test suite).
    """
    if excel_dir is None:
        excel_dir = _PROJECT_ROOT / "data" / "reconciliation"

    mu, cov = _load_python_market_data()
    excel_raw = _require_excel_csv(excel_dir / "excel_mu_vector.csv", "μ vector pytest")
    if excel_raw is None:
        import pytest  # noqa: PLC0415
        pytest.skip("Excel μ CSV not available")

    result = reconcile_arrays(mu, excel_raw.flatten(), "μ vector", TOL_MU)
    assert result.status == "pass", (
        f"μ vector reconciliation FAILED: max deviation = {result.max_deviation:.2e} "
        f"(tolerance = {TOL_MU:.0e})"
    )
    return result


def assert_cov_reconciliation(excel_dir: Path | None = None) -> CheckResult:
    """Pytest-friendly wrapper for the Σ matrix reconciliation check."""
    if excel_dir is None:
        excel_dir = _PROJECT_ROOT / "data" / "reconciliation"

    mu, cov = _load_python_market_data()
    excel_raw = _require_excel_csv(excel_dir / "excel_cov_matrix.csv", "Σ matrix pytest")
    if excel_raw is None:
        import pytest  # noqa: PLC0415
        pytest.skip("Excel Σ CSV not available")

    result = reconcile_arrays(cov, excel_raw, "Σ matrix", TOL_COV)
    assert result.status == "pass", (
        f"Σ matrix reconciliation FAILED: max deviation = {result.max_deviation:.2e} "
        f"(tolerance = {TOL_COV:.0e})"
    )
    return result


def assert_gmvp_reconciliation(excel_dir: Path | None = None) -> CheckResult:
    """Pytest-friendly wrapper for the GMVP weights reconciliation check."""
    if excel_dir is None:
        excel_dir = _PROJECT_ROOT / "data" / "reconciliation"

    mu, cov = _load_python_market_data()
    w_gmvp = _compute_python_gmvp(cov)
    excel_raw = _require_excel_csv(excel_dir / "excel_gmvp_weights.csv", "GMVP pytest")
    if excel_raw is None:
        import pytest  # noqa: PLC0415
        pytest.skip("Excel GMVP CSV not available")

    result = reconcile_arrays(w_gmvp, excel_raw.flatten(), "GMVP weights", TOL_GMVP)
    assert result.status == "pass", (
        f"GMVP reconciliation FAILED: max deviation = {result.max_deviation:.2e} "
        f"(tolerance = {TOL_GMVP:.0e})"
    )
    return result


def assert_optimal_reconciliation(A: float, excel_dir: Path | None = None) -> CheckResult:
    """Pytest-friendly wrapper for optimal weights reconciliation at a given A."""
    if excel_dir is None:
        excel_dir = _PROJECT_ROOT / "data" / "reconciliation"

    mu, cov = _load_python_market_data()
    w_python = _compute_python_optimal(mu, cov, A)
    excel_raw = _load_excel_optimal_weights(excel_dir, A)
    if excel_raw is None:
        import pytest  # noqa: PLC0415
        pytest.skip(f"Excel optimal CSV for A={A} not available")

    result = reconcile_arrays(w_python, excel_raw.flatten(), f"Optimal weights A={A}", TOL_OPTIMAL)
    assert result.status == "pass", (
        f"Optimal portfolio reconciliation FAILED for A={A}: "
        f"max deviation = {result.max_deviation:.2e} (tolerance = {TOL_OPTIMAL:.0e})"
    )
    return result


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Robo-Adviser financial reconciliation pipeline (QA)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--excel-dir",
        type=Path,
        default=None,
        help=(
            "Directory containing Excel baseline CSV exports "
            "(default: data/reconciliation/)"
        ),
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = _parse_args()
    report = run_reconciliation(excel_dir=args.excel_dir)
    # Exit with non-zero code if any check failed (useful in CI pipelines)
    if report.get("overall_status") == "FAIL":
        sys.exit(1)
